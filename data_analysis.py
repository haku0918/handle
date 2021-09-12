# -*- coding: utf-8 -*-
"""
Created on Wed Sep  8 15:46:16 2021

@author: geart
"""

from openpyxl import load_workbook
import matplotlib.pyplot as plt
import scipy.stats as stats
import pandas as pd
from statsmodels.formula.api import ols
from statsmodels.stats.anova import anova_lm
    
wb = load_workbook("C:\ibi_data\hj_normal.xlsx",data_only=True)
ws = wb['Sheet']

norm1 = []

for i in range(1,1501):
    ch = str(i)
    norm1.append(int(ws['A'+ch].value))

plt.boxplot(norm1)
plt.show()

wb2 = load_workbook("C:\ibi_data\hm_normal.xlsx",data_only=True)
ws2 = wb2['Sheet']

norm2 = []

for i in range(1,1501):
    ch = str(i)
    norm2.append(int(ws2['A'+ch].value))

plt.boxplot(norm2)
plt.show()

wb3 = load_workbook("C:\ibi_data\hj_sleepy.xlsx",data_only=True)
ws3 = wb3['Sheet']

sleepy1 = []

for i in range(1,1501):
    ch = str(i)
    sleepy1.append(int(ws3['A'+ch].value))

plt.boxplot(sleepy1)
plt.show()

wb4 = load_workbook("C:\ibi_data\hm_sleepy.xlsx",data_only=True)
ws4 = wb4['Sheet']

sleepy2 = []

for i in range(1,1501):
    ch = str(i)
    sleepy2.append(int(ws4['A'+ch].value))

plt.boxplot(sleepy2)
plt.show()
    
norm_avg=norm1
for i in range(0,1500):
    norm_avg[i]=(norm_avg[i]+norm2[i])/2

sleepy_avg=sleepy1
for i in range(0,1500):
    sleepy_avg[i]=(sleepy_avg[i]+sleepy2[i])/2


plot_data = [norm_avg, sleepy_avg]
ax = plt.boxplot(plot_data)
plt.show()

F_statistics, pval = stats.ttest_ind(norm_avg, sleepy_avg, equal_var=False)

print(F_statistics, pval)

